#!/usr/bin/env python3
"""
Fill `stock_card_no` on a Price Index CSV export from GSO's Items Export workbook.

This is the repeatable half of the Stock Card No. runbook
(docs/v1.5/STOCK_CARD_NO_RUNBOOK.md). It is offline and read-only with respect to
the portal: it takes a CSV exported from Config > Price Index plus a GSO Items
Export .xlsx, and writes a new CSV that you upload back through the same page.

Matching rule (locked in docs/v1.5/PPMP_Report_Findings.md §6):

    price-index name + unit + category  ==  export Description + Unit + Account Name

`name` + `unit` alone is ambiguous for thousands of rows, because GSO codes the
same physical item differently per account ("Battery AA" is OSAME-... as an
expense, OSAMFD-... for distribution, TREX-... under training). Category (=
Account Name) resolves nearly all of it. A name+unit-only match is used as a
fallback ONLY when the pair maps to exactly one code across the whole export.

A row is filled only when the match yields exactly ONE candidate code. Ambiguous
and unmatched rows keep whatever they already had — a wrong GSO code on an item
is worse than a blank one. Nothing is guessed, ever.

By default an existing code is preserved (blanks are filled). Pass --overwrite to
let an unambiguous match replace an existing value — use that when GSO has
reissued codes and the export is the source of truth.

Usage
-----
    python scripts/stock_card_no_merge.py \
        --price-index ~/Downloads/price-index.csv \
        --items-export ~/Downloads/Items_Export_2026-07-24.xlsx \
        --out ~/Downloads/price-index-with-codes.csv \
        --report ~/Downloads/stock-card-unresolved.csv

Requires: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

# The Price Index CSV contract (PriceIndexService.CsvHeaders). stock_card_no is
# last on purpose: a pre-v1.5 file without it still imports, and its absence means
# "leave alone" rather than "clear".
PRICE_INDEX_HEADERS = ["name", "unit", "unit_price", "category", "is_active", "days_enabled", "stock_card_no"]

# Columns in GSO's Items Export workbook, as of the 2026-07-24 export.
EXPORT_CODE = "Item Code"
EXPORT_NAME = "Description"
EXPORT_UNIT = "Unit"
EXPORT_ACCOUNT = "Account Name"

# Excel exports carry literal _x000d_ escapes and embedded newlines in Description.
_XML_CR = re.compile(r"_x000d_", re.IGNORECASE)
_WHITESPACE = re.compile(r"\s+")

# 1,040 codes in the 2026-07-24 export use an EN DASH (U+2013) where every other code uses a
# hyphen — "RAM–BAOS-1434419431" vs "RAM-BAOS-1434419431". They look identical on screen but
# would not match a typed search in the config UI, so codes are normalized to ASCII hyphens on
# the way in. Do not remove this without checking a fresh export first.
_DASHES = str.maketrans({"–": "-", "—": "-", "−": "-", "‐": "-", "‑": "-"})


def clean_code(value: object) -> str:
    """Normalize a GSO item code: trim, collapse inner whitespace, unify dashes to ASCII hyphen."""
    if value is None:
        return ""
    return _WHITESPACE.sub("", str(value).translate(_DASHES)).strip()


def norm(value: object) -> str:
    """Normalize a cell for matching: strip Excel CR escapes, collapse whitespace, casefold."""
    if value is None:
        return ""
    text = _XML_CR.sub(" ", str(value))
    return _WHITESPACE.sub(" ", text).strip().casefold()


def load_export(path: Path) -> tuple[dict, dict]:
    """Return (by_name_unit_account, by_name_unit) — each mapping a key to a set of item codes."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - operator-facing
        sys.exit("openpyxl is required:  pip install openpyxl")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        sys.exit(f"{path} is empty.")

    missing = [c for c in (EXPORT_CODE, EXPORT_NAME, EXPORT_UNIT, EXPORT_ACCOUNT) if c not in header]
    if missing:
        sys.exit(
            f"{path.name} is missing expected column(s): {', '.join(missing)}.\n"
            f"Found: {', '.join(header)}\n"
            "If GSO renamed the columns, update the EXPORT_* constants at the top of this script."
        )

    idx = {name: header.index(name) for name in (EXPORT_CODE, EXPORT_NAME, EXPORT_UNIT, EXPORT_ACCOUNT)}
    by_three: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    by_two: dict[tuple[str, str], set[str]] = defaultdict(set)

    for row in rows:
        code = clean_code(row[idx[EXPORT_CODE]])
        if not code:
            continue
        name, unit, account = norm(row[idx[EXPORT_NAME]]), norm(row[idx[EXPORT_UNIT]]), norm(row[idx[EXPORT_ACCOUNT]])
        if not name or not unit:
            continue
        by_three[(name, unit, account)].add(code)
        by_two[(name, unit)].add(code)

    return by_three, by_two


def read_price_index(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            sys.exit(f"{path} is empty.")
        if "name" not in reader.fieldnames or "unit" not in reader.fieldnames:
            sys.exit(
                f"{path.name} does not look like a Price Index export "
                f"(expected headers: {', '.join(PRICE_INDEX_HEADERS)})."
            )
        return [dict(row) for row in reader]


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill stock_card_no on a Price Index CSV from a GSO Items Export.")
    parser.add_argument("--price-index", required=True, type=Path, help="CSV exported from Config > Price Index")
    parser.add_argument("--items-export", required=True, type=Path, help="GSO Items Export .xlsx")
    parser.add_argument("--out", required=True, type=Path, help="CSV to write (upload this back to the portal)")
    parser.add_argument("--report", type=Path, help="Optional CSV listing rows left unresolved")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Let an unambiguous match replace an existing stock_card_no (default: fill blanks only)",
    )
    args = parser.parse_args()

    by_three, by_two = load_export(args.items_export)
    rows = read_price_index(args.price_index)

    filled_three = filled_two = ambiguous = no_match = kept_existing = unchanged_blank = 0
    unresolved: list[dict[str, str]] = []

    for row in rows:
        existing = (row.get("stock_card_no") or "").strip()
        key3 = (norm(row.get("name")), norm(row.get("unit")), norm(row.get("category")))
        key2 = (key3[0], key3[1])

        candidates = by_three.get(key3) or set()
        source = "name+unit+category"
        if not candidates:
            candidates = by_two.get(key2) or set()
            source = "name+unit"

        code = next(iter(candidates)) if len(candidates) == 1 else None
        reason = "" if code else ("ambiguous" if len(candidates) > 1 else "no match in export")

        if code and existing and not args.overwrite:
            kept_existing += 1
            continue
        if code:
            row["stock_card_no"] = code
            if source == "name+unit+category":
                filled_three += 1
            else:
                filled_two += 1
            continue

        # Unresolved: preserve whatever was there. Never blank a curated value.
        row["stock_card_no"] = existing
        if len(candidates) > 1:
            ambiguous += 1
        else:
            no_match += 1
        if not existing:
            unchanged_blank += 1
        unresolved.append(
            {
                "name": row.get("name", ""),
                "unit": row.get("unit", ""),
                "category": row.get("category", ""),
                "existing_stock_card_no": existing,
                "reason": reason,
                "candidate_codes": " | ".join(sorted(candidates)),
            }
        )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=PRICE_INDEX_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in PRICE_INDEX_HEADERS})

    if args.report:
        with args.report.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["name", "unit", "category", "existing_stock_card_no", "reason", "candidate_codes"],
            )
            writer.writeheader()
            writer.writerows(unresolved)

    total_filled = filled_three + filled_two
    print(f"Price index rows read      : {len(rows)}")
    print(f"Filled (name+unit+category): {filled_three}")
    print(f"Filled (name+unit only)    : {filled_two}")
    print(f"Existing code kept         : {kept_existing}" + ("" if args.overwrite else "  (use --overwrite to replace)"))
    print(f"Ambiguous - left as-is     : {ambiguous}")
    print(f"No match  - left as-is     : {no_match}")
    print(f"Still blank after this run : {unchanged_blank}")
    print(f"\nWrote {args.out}  ({total_filled} code(s) written)")
    if args.report:
        print(f"Wrote {args.report}  ({len(unresolved)} unresolved row(s))")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
