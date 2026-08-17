"""Build the v1.8.0 Monday question workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

FONT = "Arial"
GREEN = "196638"      # PPDO green-700
GREEN_LT = "D4EDDE"
RED_LT = "FEE2E2"
YELLOW = "FFF9C4"
GREY_LT = "F1F5F9"
BORDER = Border(*[Side(style="thin", color="CBD5E1")] * 4)

HEADERS = [
    ("ID", 7),
    ("Priority", 11),
    ("Who answers", 14),
    ("Topic", 18),
    ("Question", 46),
    ("Why it matters", 42),
    ("Example / options", 40),
    ("Our recommendation", 40),
    ("ANSWER", 30),
    ("Answered by / date", 16),
    ("Developer notes (technical)", 52),
]

B = "BLOCKER"
S = ""

ROWS = [
    # --- A. PBO -------------------------------------------------------------
    ("A1", B, "PBO / PPDC", "Budget ceiling",
     "Does the AIP draw from the same budget ceiling the WFP already uses, or its own separate one?",
     "The same peso could be counted twice — once when planned in the AIP, again when detailed in the WFP. There would be no error message; the numbers would just be wrong.",
     "Same pot: the ceiling limits the AIP, and the WFP is limited by its AIP activity only.\nSeparate pots: the AIP gets its own ceiling figure.",
     "Same pot, with the AIP as the limit. The AIP is the plan; the WFP details it. But this changes WFP behaviour that already works.",
     "", "",
     "WfpCeilingService runs two checks per expenditure: against AipActivity.Total (x1000) and against DivisionAllocation. Adding allocation->AIP closes a triangle. If 'same pot', the division-allocation check in WFP becomes redundant and MUST be removed or every peso is deducted twice. Decides whether we add AipDivisionAllocationLedger or make WfpDivisionAllocationLedger polymorphic (sourceType, sourceId)."),

    ("A2a", B, "PBO", "Rounding",
     "On one line, PS + MOOE + CO must equal the Total. If each is rounded up on its own, the row stops adding up. Which figure wins?",
     "This shows on EVERY printed line, so anyone checking a single row sees it immediately.",
     "Exact:   PS 100,100 | MOOE 200,100 | CO 300,100 | Total 600,300\nRounded: PS 101 | MOOE 201 | CO 301 | Total 601\nBut 101+201+301 = 603, not 601.",
     "Make the row total the sum of the rounded components (603). Combined with A3, the document then balances both across and down.",
     "", "",
     "Applies to the AIP form's PS/MOOE/CO/Total columns. Up to 2 thousand-units of drift per row. Implement as: round components first, derive every total from rounded inputs. Same helper used by UI and Excel export so they cannot diverge."),

    ("A2b", S, "PBO", "Rounding",
     "Any amount under P1,000 will print as 1. Is that acceptable?",
     "A P500 line printing as '1' overstates it, but that follows from the never-understate rule.",
     "P0.00 -> 0\nP0.01 -> 1\nP500 -> 1\nP999.99 -> 1",
     "Accept. Consistent with the agreed rule, and AIP lines are rarely this small.",
     "", "",
     "Math.Ceiling(amount / 1000m). Zero must stay zero - the rule triggers on a remainder ABOVE zero, so an empty line does not become 1."),

    ("A2c", S, "PBO / PPDC", "Rounding",
     "Can an amount ever be negative — for example a reduction in an amendment or supplemental budget?",
     "'Round up' on a negative number makes the deduction SMALLER, which is the opposite of the conservative intent.",
     "-1,234,200 rounds to -1,234 (a smaller deduction)\nThe intent would want -1,235.",
     "If negatives are possible, 'round up' should mean round the SIZE up (away from zero), not numerically up.",
     "", "",
     "Math.Ceiling on a negative rounds toward zero. Would need Math.Sign(x) * Math.Ceiling(Math.Abs(x)/1000m). Moot for a first AIP (a plan of appropriations); becomes real with the amendment/supplemental flow (RAL-78)."),

    ("A2d", S, "PBO", "Rounding",
     "Should the budget ceiling be checked against the rounded figures or the exact ones?",
     "They can disagree — the system could say an office is within its ceiling while the printed AIP reads as over it.",
     "Exact total:   P49,999,600 (inside a P50M ceiling)\nPrinted total: P50,020,000 (reads as P20,000 over)",
     "Enforce against the same rounded figures the document prints, so the system and the paper can never contradict each other.",
     "", "",
     "Pairs with A3 - same underlying choice about which figures are authoritative. If we enforce on exact but print rounded, expect support calls from reviewers reading the printed document."),

    ("A2e", S, "PBO", "Rounding",
     "In the Excel file we send you, should the cells contain the rounded thousands or the exact pesos?",
     "If you re-add a column in Excel, the result should match the printed total exactly.",
     "Rounded: cell holds 1235\nExact:   cell holds 1234567.89 formatted to look like 1,235",
     "Rounded thousands, as real numbers, with the sheet labelled '(In Thousand Pesos)'.",
     "", "",
     "If we write exact pesos with display formatting, the recipient's own SUM() disagrees with our printed total. Write the rounded value; keep exact values in the DB only."),

    ("A3", B, "PBO", "Rounding",
     "Going down a column, must the printed figures add up exactly to the printed total?",
     "Because everything rounds up, every row is overstated and nothing cancels out. The gap grows with the number of rows and is always in the same direction.",
     "10 rows  -> column exceeds total by ~5 thousand\n100 rows -> ~50 thousand\n500 rows -> ~250 thousand",
     "Round each row first, then add the rounded rows for every subtotal and total. Still never understates, and the document is internally consistent.",
     "", "",
     "Two options: (a) sum exact then round the total - accurate but the printed column visibly does not add up; (b) round rows then sum - column adds up, total sits above the true sum. Under ceiling rounding option (a) drifts ~500 pesos per row, one-directional."),

    ("A4", S, "PBO", "Ceiling behaviour",
     "If an office goes over its ceiling, should the system block them from saving, or just warn?",
     "A hard block while typing is frustrating on a document built over weeks; a warning risks over-budget submissions.",
     "Block on save / Block only at submit / Warn only",
     "Block at submit, warn while typing. Also near-necessary if people work offline, where a live block cannot be evaluated.",
     "", "",
     "Offline entry (PWA) cannot evaluate a hard block at typing time without cached allocations. Submit-time validation is the natural choke point and matches the completeness-check idea."),

    ("A5", S, "PBO", "Allocation",
     "Must an office's division allocations add up to exactly its ceiling, or can they total less? And what happens if PBO lowers a ceiling after divisions have already encoded against it?",
     "Without a rule, an office could allocate more to divisions than it actually has.",
     "Must equal / May be less / May exceed",
     "Allocations must not exceed the ceiling. If a ceiling is lowered below what is already encoded, flag the affected divisions rather than silently invalidating work.",
     "", "",
     "No such check exists today. Needs a validation in AllocationService plus a recalculation path when a ceiling changes after encoding has begun."),

    ("A6a", S, "PBO", "Confirmation",
     "Confirm: only the General Fund has a ceiling. GAD, 20% Development Fund, LDRRF and Trust Fund have none.",
     "This is what we understood from the discussion; confirming before it is built in.",
     "Yes / No",
     "-",
     "", "",
     "Ceiling check filters on funding_source_id = General Fund. Resolves whiteboard item W1."),

    ("A6b", S, "PBO", "Confirmation",
     "Confirm: Personal Services is excluded from the ceiling even when it is General Fund. So if one expense is part PS and part MOOE, only the MOOE part counts against the ceiling.",
     "PS is an expense class, not a fund source, so the exclusion works differently from the others.",
     "Yes / No",
     "-",
     "", "",
     "The check must sum only the (mooe + co) portions of General-Fund lines, NOT whole General-Fund lines. Easy to implement the wrong one; the difference is invisible until audited."),

    ("A6c", S, "PBO", "Confirmation",
     "Confirm: amounts are entered in full pesos (e.g. 1,234,567.89) and only DISPLAYED in thousands.",
     "The exact figure stays in the system; thousands is a presentation convention.",
     "Yes / No",
     "-",
     "", "",
     "IMPORTANT: AIP amounts are currently stored IN THOUSANDS (the .xlsm is written that way). This inverts the storage unit. See D1 - three x1000 conversions in WfpCeilingService must be removed in the same change."),

    # --- B. PPDC ------------------------------------------------------------
    ("B1", B, "PPDC", "Review workflow",
     "Can a reviewer send work back for correction, and can they leave comments? If comments — on the whole submission, or on a specific program / project / activity / expense line?",
     "The requirements describe submitting and approving, but not what happens when a reviewer is not satisfied.",
     "Approve only / Approve or return\nComments: whole submission / per item",
     "Approve or return, with per-item comments. Per-item is much more useful to the encoder, though it is more work to build.",
     "", "",
     "Adds a Returned state plus a comment thread. Per-node comments need a comments table keyed to the hierarchy node. Whole-submission comments are one text column. PlanningStatus is currently just Draft/Final/Archived (string constants, not a DB enum) so extending is cheap."),

    ("B2", B, "PPDC", "Entry",
     "Can an office add a program that is not in its approved LDIP?",
     "Programs load from the LDIP. If something is missing, offices need to know whether they can add it themselves.",
     "Yes, freely / Yes, with approval / No, LDIP is the closed list",
     "-",
     "", "",
     "Decides whether program creation exists in the AIP UI at all. seedAipProgramsFromLdip (RAL-181) already loads them. Related: who assigns programs to divisions for non-PPDO offices?"),

    ("B3", S, "PPDC", "Review workflow",
     "Once work is submitted, is it locked? Locked to the encoder only, or to the reviewer too? Can the reviewer edit figures, or only approve and comment? If work is returned and resubmitted, should the reviewer see what changed?",
     "Determines who can change numbers after submission, and whether a reviewer can spot what was altered.",
     "Locked to encoder / Locked to all\nReviewer: read-only / can edit",
     "Reviewer read-only on content (as previously agreed), and show what changed on resubmission.",
     "", "",
     "Reviewer read-only is the first SUBTRACTIVE permission in the codebase - every existing flag only grants. Cannot use the existing ConfigHttp.AuthorizeAsync idiom; needs a deny-guard on every write endpoint. 'What changed' needs a diff or a snapshot."),

    ("B4", S, "PPDC", "Review workflow",
     "Is PPDO's own internal review a real approval gate, or just a view? In other words, can PPDO's divisions reach the consolidated document without the PPDO reviewer approving first?",
     "Determines whether PPDO has the same two-step flow as the other offices.",
     "PPDO division encoder -> PPDO reviewer -> Consolidated -> LFC\nOR: PPDO divisions -> Consolidated directly",
     "-",
     "", "",
     "Decides whether PPDO gets the same submit/approve transitions as other offices, or whether its roll-up is implicit. Affects the state machine and the consolidation query."),

    ("B5", S, "PPDC / LFC", "Review workflow",
     "When the LFC reviews the consolidated AIP, can they return ONE office's work, or only approve/return the whole document? And if one office is returned, must its reviewer submit again?",
     "Determines whether one office's correction holds up everyone else.",
     "Per office / Whole document only",
     "Per office. Otherwise a single correction blocks all 20 offices.",
     "", "",
     "Decides whether consolidation is a live view over approved office records (per-office return is easy) or a snapshot (per-office return means re-consolidating)."),

    ("B6", S, "PPDC", "Users",
     "Who are the LFC users — names, or at least how many? And confirm they are meant to see every office's budget figures.",
     "LFC would be the first role in the system that can see across all offices; everyone else sees only their own.",
     "-",
     "-",
     "", "",
     "First cross-office permission. Its resolution must BYPASS OfficeScope rather than combine with it, and must not inherit the reviewer flag's write-denial semantics. Needs its own flag, not reviewer + all offices."),

    ("B7", S, "PPDC", "Deadline",
     "Should there be a submission cut-off date per fiscal year, after which offices can no longer submit? And a page showing which offices have submitted and which have not?",
     "With ~20 offices, chasing submissions manually is the alternative.",
     "Yes / No",
     "Yes. It is also the natural point to enforce the ceiling check (A4).",
     "", "",
     "Small: a deadline column per FY plus a readiness view. Gives submit-time validation a natural hard-block point."),

    ("B8", S, "PPDC", "Clarification",
     "On the whiteboard there was 'P5M' written next to Program / Project. What was that?",
     "Could not tell from the photo whether it is a threshold, a ceiling, or just an example figure.",
     "Threshold / Ceiling / Example only",
     "-",
     "", "",
     "Whiteboard item W9. If it is a project-level threshold it may imply an extra validation or an approval tier."),

    # --- C. Consuming offices ----------------------------------------------
    ("C1", S, "PBO, PACCO,\nPTO, GSO", "Data exports",
     "What do you need out of the AIP, in what format, and how often? Please share a filled-in copy of whatever you use today rather than a description.",
     "A real example removes the guesswork about columns and layout.",
     "Excel / CSV / direct connection\nOnce on approval / continuously refreshed",
     "Build one complete dataset (one row per expense line) and give each office a filtered view of it, rather than five separate reports.",
     "", "",
     "Five bespoke reports will drift. One canonical dataset + column selection will not. NOTE the v1.5 lesson: the province's 'WFP template' turned out to be a filled sample, which cost real time - insist on a real file."),

    ("C2", S, "GSO", "Data exports",
     "For GSO specifically — would a live connection into your system be more useful than a file?",
     "A file goes stale as soon as it is generated; a live connection always shows current data.",
     "File / Live connection",
     "Ask before building a file. A read-only API for GSO has already been designed.",
     "", "",
     "docs/External_AIP_API_Contract.md already exists in the backlog for exactly this. May supersede a GSO file export entirely."),

    # --- D. Ralph -----------------------------------------------------------
    ("D1", B, "Ralph", "Existing data",
     "What happens to the 2027 AIP when the new AIP goes live — is it converted to the new format, or kept as-is with the new format applying only from 2028? And does the Excel upload stay?",
     "2027 is live data and the WFP is built on it, so it cannot simply be replaced.",
     "Convert 2027 / Keep 2027 as-is\nExcel upload: keep / retire",
     "Convert 2027. Leaving two different units in the same column is a bug waiting to happen.",
     "", "",
     "CRITICAL: three x1000 conversions in WfpCeilingService (lines ~60, ~106, ~220) convert AIP thousands to WFP pesos. When AIP moves to peso storage these MUST be removed in the same change, or every WFP ceiling check silently becomes 1000x too permissive - no error, no stack trace. Also: wfp_activities.aip_activity_id is FK-Restrict onto AIP activities. Add a test pinning the AIP/WFP comparison."),

    ("D2", S, "Ralph", "Printable AIP",
     "The printable AIP — should it match the same official form layout that the Excel upload reads?",
     "If so, the structure is already known and does not need to be worked out from scratch.",
     "Same as the imported .xlsm form / Something different",
     "Same layout. The importer already knows this structure.",
     "", "",
     "AipXlsmParser already encodes the GENERAL_/SOCIAL_/ECONOMIC_/OTHERS_ sheet structure, so the exporter can be derived from it. Follow the v1.4.4/v1.5 rule: build the sheet programmatically from a documented style catalogue, do NOT clone rows out of a reference workbook."),

    ("D3", S, "Ralph", "Offline",
     "Will offline data entry happen on office-issued machines, or on personal / shared laptops?",
     "Decides how much of the login session we are willing to leave on the device.",
     "Office-issued / Personal or shared",
     "If shared: allow offline work without a saved login, but require a real login to upload.",
     "", "",
     "Today the access token is in-memory only and never persisted, so a stolen laptop yields nothing. Persisting a session inverts that. Recommend IndexedDB for drafts (not localStorage - synchronous, ~5MB, blocks the UI mid-typing), plus 'sign out and clear local work' and an auto-wipe after N days."),

    ("D4", S, "Ralph / PPDC", "Concurrency",
     "Is it one encoder per office, or could two people encode the same office's AIP at the same time?",
     "If two, we need a rule for what happens when they edit the same activity.",
     "One encoder / Multiple encoders",
     "If multiple: a soft lock or a 'changed by someone else' warning. Currently the last save silently wins.",
     "", "",
     "Also affects the offline merge story - two encoders working offline on the same office produce a genuine merge conflict at upload."),
]

SETTLED = [
    ("Multiple fund sources on one activity",
     "One fund source per expense line. An activity using several funds simply has several lines.",
     "Ralph, 14 Aug 2026",
     "aip_expenditures gets a single funding_source_id. The simpler option, and it can express everything the alternative could."),
    ("Rounding to thousands",
     "Always round UP whenever there is any remainder above zero. 1,234,200 -> 1,235. Zero stays zero.",
     "Ralph, 14 Aug 2026",
     "Math.Ceiling(amount / 1000m). Rationale: better to state slightly more than slightly less. See A2a-A2e for the follow-on details."),
    ("Which funds have a ceiling",
     "General Fund only. GAD, 20% DF, LDRRF and Trust Fund have none. Personal Services is exempt.",
     "Ralph, 14 Aug 2026",
     "Resolves whiteboard item W1, which was NOT a permission rule as first read - it is the ceiling exemption list."),
    ("Printable AIP form",
     "In scope for v1.8.0.",
     "Ralph, 14 Aug 2026",
     "The largest single deliverable missing from the original draft. See D2 for the layout question."),
    ("Password reset",
     "Self-service: the user answers their own recovery question and gets a temporary password. No admin involvement.",
     "Ralph, 14 Aug 2026",
     "Option B. Recovery answer hashed like a password; attempt lockout; identical response for wrong answer and unknown username; audit entry plus a 'your password was reset' notice. No AIP dependency - can be built first."),
    ("Suggested additions",
     "Accepted: submission deadline, in-app notifications, amendment readiness, carry-forward from prior year, completeness checks before submit, approval snapshot, concurrent-edit guard.",
     "Ralph, 14 Aug 2026",
     "Detailed in AIP_Requirements_Review.md section 9."),
]


def style_header(ws, headers, row=1):
    for idx, (title, width) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=title)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 30


wb = Workbook()

# ── Sheet 1 — Read Me ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 86

def rm(row, label, value, bold_label=True, label_size=10):
    if label:
        c = ws.cell(row=row, column=2, value=label)
        c.font = Font(name=FONT, size=label_size, bold=bold_label)
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if value:
        c = ws.cell(row=row, column=3, value=value)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)

t = ws.cell(row=2, column=2, value="PPDO Portal v1.8.0 — Questions for Monday")
t.font = Font(name=FONT, size=16, bold=True, color=GREEN)
sub = ws.cell(row=3, column=2, value="AIP redesign, allocation, password reset and offline data entry")
sub.font = Font(name=FONT, size=11, italic=True, color="475569")

rm(5, "What this is", "The open questions from the v1.8.0 requirements review. Each one is written for the person answering it, not for the developer asking.")
rm(6, "How to use it", "Go to the 'Questions' tab. Write answers in the yellow ANSWER column. Nothing else needs filling in.")
rm(7, "If time is short", "Answer the rows marked BLOCKER first. Development cannot start on those areas until they are settled.")
rm(9, "Tabs", "Read Me — this page\nQuestions — the list, grouped by who answers\nAlready Settled — decisions already made, listed so they are not reopened by mistake")
ws.row_dimensions[9].height = 46

rm(11, "Who answers what", "PBO — the money rules: ceilings, rounding, what the printed figures must add up to\nPPDC / PPDO leadership — the review and approval workflow\nPBO, PACCO, PTO, GSO — what data they each need out of the AIP\nRalph — internal technical decisions, no meeting needed")
ws.row_dimensions[11].height = 60

rm(13, "Columns", "Question, Why it matters, Example / options and Our recommendation are written in plain language.\nDeveloper notes (last column) is technical detail for the build — it can be ignored in the meeting.")
ws.row_dimensions[13].height = 32

# Live status counters
c = ws.cell(row=15, column=2, value="Progress")
c.font = Font(name=FONT, size=10, bold=True)
for i, (label, formula) in enumerate([
    ("Total questions", '=COUNTA(Questions!A2:A100)'),
    ("Blockers", '=COUNTIF(Questions!B2:B100,"BLOCKER")'),
    ("Answered", '=COUNTA(Questions!I2:I100)'),
    ("Blockers still open", '=COUNTIFS(Questions!B2:B100,"BLOCKER",Questions!I2:I100,"")'),
]):
    r = 16 + i
    lc = ws.cell(row=r, column=2, value=label)
    lc.font = Font(name=FONT, size=10)
    lc.alignment = Alignment(indent=1)
    vc = ws.cell(row=r, column=3, value=formula)
    vc.font = Font(name=FONT, size=10, bold=True)
    vc.alignment = Alignment(horizontal="left")

rm(22, "Prepared", "14 August 2026. Full technical reasoning: docs/v1.8/AIP_Requirements_Review.md in the repository.")

# ── Sheet 2 — Questions ────────────────────────────────────────────────────
qs = wb.create_sheet("Questions")
style_header(qs, HEADERS)
qs.freeze_panes = "E2"
qs.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

for r, row in enumerate(ROWS, start=2):
    blocker = row[1] == B
    for i, val in enumerate(row, start=1):
        c = qs.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=10,
                      bold=(i in (1, 2) and blocker),
                      color="991B1B" if (i == 2 and blocker) else "000000")
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if i in (1, 2) else "left")
        c.border = BORDER
        if i == 2 and blocker:
            c.fill = PatternFill("solid", fgColor=RED_LT)
        elif i == 9:                       # ANSWER column
            c.fill = PatternFill("solid", fgColor=YELLOW)
        elif i == 10:
            c.fill = PatternFill("solid", fgColor=YELLOW)
        elif i == 11:                      # developer notes
            c.fill = PatternFill("solid", fgColor=GREY_LT)
            c.font = Font(name=FONT, size=9, color="334155")

qs.page_setup.orientation = "landscape"
qs.page_setup.fitToWidth = 1
qs.page_setup.fitToHeight = 0
qs.sheet_properties.pageSetUpPr.fitToPage = True
qs.print_title_rows = "1:1"
qs.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

# ── Sheet 3 — Already Settled ──────────────────────────────────────────────
st = wb.create_sheet("Already Settled")
SET_HEADERS = [("Topic", 34), ("Decision", 62), ("Decided by", 18), ("Developer notes (technical)", 66)]
style_header(st, SET_HEADERS)
st.freeze_panes = "A2"

for r, row in enumerate(SETTLED, start=2):
    for i, val in enumerate(row, start=1):
        c = st.cell(row=r, column=i, value=val)
        c.font = Font(name=FONT, size=10) if i != 4 else Font(name=FONT, size=9, color="334155")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER
        if i == 4:
            c.fill = PatternFill("solid", fgColor=GREY_LT)
        elif i == 2:
            c.fill = PatternFill("solid", fgColor=GREEN_LT)

st.page_setup.orientation = "landscape"
st.sheet_properties.pageSetUpPr.fitToPage = True
st.page_setup.fitToWidth = 1
st.page_setup.fitToHeight = 0

out = "/home/user/ppdo-portal/docs/v1.8/v1.8.0_Questions_for_Monday.xlsx"
wb.save(out)
print("wrote", out, "|", len(ROWS), "questions,", sum(1 for r in ROWS if r[1] == B), "blockers")
